class column_config(object):
    def __init__(self):
        self.field = None
        self.caption = None
        self.is_hidden = False
    def init_data(self,*args,**kwargs):
        if type(args) is tuple:
            self.field = args[0]
            self.caption = args[1]
            self.depth_fields=self.field.split('.')
            self.depth_count=self.depth_fields.__len__()
            if args.__len__()>2:
                self.is_hidden =args[2]
        return self
class config(object):
    def __init__(self):
        self.columns=[]
        self.wb = None
        self.ws = None
    def init_data(self,*args,**kwargs):
        for x in args:
            col = column_config()
            self.columns.append(
                col.init_data(*x)
            )
        return self
    @property
    def workbook(self):
        return self.wb
    def save(self,path_to_file):
        self.wb.save(path_to_file)
        return self
    def exract_item(self,item):
        ret_list =[]
        for col in self.columns:
            ret = item
            for i in range(0,col.depth_count-1):
                ret=ret.get(col.depth_fields[i],{})
            ret_list.append(ret.get(col.depth_fields[col.depth_count-1],None))
        return ret_list


    def unwind_data(self,items):
        for x in items:
            yield self.exract_item(x)

    def fetch_data(self,items):
        from openpyxl import Workbook
        from openpyxl import utils
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "data"
        caption_list = [x.caption for x in self.columns]
        field_list = [x.field for x in self.columns]
        self.ws.append(caption_list)
        col_index = 1
        for x in field_list:
            self.wb.create_named_range(
                x,
                self.ws,
                "$" + utils.get_column_letter(col_index) + ":$" +
                utils.get_column_letter(col_index)
            )
            col_index = col_index + 1

        for row in self.unwind_data(items):
            self.ws.append(row)
        return self


def create(*args,**kwargs):
    ret = config()
    ret.init_data(*args,**kwargs)
    return ret




