def do_export(lan,app_name,schema,source,coll,cols_caption=None):
    from . import language

    if not hasattr(coll,"get_list"):
        raise (Exception("It looks like you export data from none-datasource. Datasource must have 'get_list' method that return list"))
    selected_fields=coll.get_selected_fields()
    col_of_fields = language.get_columns(lan, app_name, schema, source,selected_fields)
    col_of_fields =sorted(col_of_fields,lambda x,y: x["display_index"]-y["display_index"])

    projector={
        "_id":0
    }

    for x in col_of_fields:
        projector.update({
            x['field']:1
        })

    caption_list = [ x.get("caption") for x in col_of_fields]
    field_list =[ x.get("field") for x in col_of_fields]

    coll.project(projector)
    lst = coll.get_list()
    from openpyxl import Workbook
    from openpyxl import utils
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(caption_list)
    col_index = 1
    for x in field_list:
        wb.create_named_range(
            x,
            ws,
            "$"+utils.get_column_letter(col_index)+":$"+
            utils.get_column_letter(col_index)
        )
    # for x in header_attr:
    #     wb.create_named_range(x, ws,
    #                           "$" + utils.get_column_letter(col_index) + ":$" + utils.get_column_letter(col_index))


    # col_index = col_index + 1
    for x in lst:
        items =[]
        for f in field_list:
            items.append(x.get(f,None))
        ws.append(items)
    from openpyxl.writer.excel import save_virtual_workbook
    from django.http import HttpResponse
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename= "{}"'.format(source+".xlsx")
    return response

